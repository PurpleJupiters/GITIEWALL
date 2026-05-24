"""
SoundCloud Following Fetcher — Agent WALL (wall-0)
Fetches all followed artists with: name, city, country, followers,
join date, tracks, profile URL. Saves to Excel + CSV.
"""

import requests
import browser_cookie3
import openpyxl
import csv
import json
import time
import sys
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

USER_ID = 196753
CLIENT_ID = "tUy37JutyVy6r6JSMLnScSmBwA5DoTXE"
BASE_URL = "https://api-v2.soundcloud.com"
OUTPUT_XLSX = r"E:\SunoMaster\scripts\soundcloud_following.xlsx"
OUTPUT_CSV  = r"E:\SunoMaster\scripts\soundcloud_following.csv"
OUTPUT_JSON = r"E:\SunoMaster\scripts\soundcloud_following.json"


def get_chrome_cookies():
    import os
    auth_file = r"C:\Users\equat\Downloads\sc_auth.json"
    try:
        with open(auth_file, "r") as f:
            auth_data = json.load(f)
        auth_token = auth_data.get("auth", "")
        session = requests.Session()
        session.headers.update({
            "Authorization": auth_token,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
            "Origin": "https://soundcloud.com",
            "Referer": "https://soundcloud.com/wall-0/following",
        })
        print(f"  Auth token loaded from {auth_file}")
        return session
    except Exception as e:
        print(f"Auth file error: {e}")
        return None


def fetch_all_followings(session):
    from urllib.parse import urlparse, parse_qs
    all_users = []
    seen = set()
    offset = None
    page = 0

    while True:
        page += 1
        if offset:
            url = f"{BASE_URL}/users/{USER_ID}/followings?limit=200&offset={offset}&client_id={CLIENT_ID}&app_version=1779379648&app_locale=en"
        else:
            url = f"{BASE_URL}/users/{USER_ID}/followings?limit=200&client_id={CLIENT_ID}&app_version=1779379648&app_locale=en"

        print(f"  Page {page}: offset={offset} ...")
        try:
            resp = session.get(url, timeout=15, verify=False)
            if resp.status_code != 200:
                print(f"  ERROR: status {resp.status_code} on page {page}")
                break
            data = resp.json()
        except Exception as e:
            print(f"  ERROR fetching page {page}: {e}")
            break

        collection = data.get("collection", [])
        added = 0
        for user in collection:
            uid = user.get("id")
            if uid and uid not in seen:
                seen.add(uid)
                all_users.append(user)
                added += 1

        print(f"    Got {len(collection)} users ({added} new). Total: {len(all_users)}")

        # Extract offset from next_href without using the signed URL directly
        next_href = data.get("next_href")
        if not next_href or len(collection) == 0:
            break

        parsed = urlparse(next_href)
        params = parse_qs(parsed.query)
        next_offset = params.get("offset", [None])[0]

        if not next_offset or next_offset == offset:
            break
        offset = next_offset
        time.sleep(0.3)

    return all_users


def parse_user(user):
    joined_raw = user.get("created_at") or ""
    if joined_raw:
        try:
            dt = datetime.strptime(joined_raw[:10], "%Y-%m-%d")
            joined = dt.strftime("%Y-%m-%d")
            years_on_sc = round((datetime.now() - dt).days / 365.25, 1)
        except:
            joined = joined_raw[:10]
            years_on_sc = ""
    else:
        joined = ""
        years_on_sc = ""

    return {
        "Username":       user.get("username", ""),
        "Full Name":      user.get("full_name", ""),
        "City":           user.get("city", ""),
        "Country":        user.get("country_code", ""),
        "Followers":      user.get("followers_count", 0),
        "Following":      user.get("followings_count", 0),
        "Tracks":         user.get("track_count", 0),
        "Joined SC":      joined,
        "Years on SC":    years_on_sc,
        "Verified":       "Yes" if user.get("verified") else "No",
        "Pro":            user.get("creator_subscription", {}).get("product", {}).get("id", ""),
        "Profile URL":    user.get("permalink_url", ""),
        "Unfollow?":      "TBD",  # set after ranking
    }


def save_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoundCloud Following"

    headers = list(rows[0].keys()) if rows else []

    # Header style
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="FF6B35", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    red_fill   = PatternFill("solid", fgColor="FFD7D7")
    green_fill = PatternFill("solid", fgColor="D7FFD7")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            val = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")
        # Highlight rows: red = unfollow candidate, green = keep
        highlight = red_fill if row.get("Unfollow?") == "UNFOLLOW" else green_fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = highlight

    # Column widths
    col_widths = {
        "Rank": 6, "Username": 25, "Full Name": 22, "City": 18, "Country": 8,
        "Followers": 12, "Following": 12, "Tracks": 8,
        "Joined SC": 12, "Years on SC": 12, "Verified": 9, "Pro": 14,
        "Profile URL": 45, "Unfollow?": 12
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(rows)
    to_unfollow = sum(1 for r in rows if r.get("Unfollow?") == "YES")
    to_keep = total - to_unfollow

    summary = [
        ("Total following", total),
        ("Keep (top 800 by followers)", to_keep),
        ("Unfollow candidates (rank 801+)", to_unfollow),
        ("% reduction", f"{round(to_unfollow/total*100, 1)}%" if total else "0%"),
        ("Rank 800 follower threshold", rows[799]["Followers"] if len(rows) >= 800 else "N/A"),
        ("NOTE: Review list — override UNFOLLOW to 'keep' for personal contacts", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r_idx, (label, value) in enumerate(summary, 1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r_idx, column=2, value=value)

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 20

    wb.save(OUTPUT_XLSX)
    print(f"  Excel saved: {OUTPUT_XLSX}")


def save_csv(rows):
    if not rows:
        return
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"  CSV saved: {OUTPUT_CSV}")


def main():
    print("=== SoundCloud Following Fetcher ===")
    print(f"Profile: wall-0 (user ID {USER_ID})")
    print("Getting Chrome cookies...")

    session = get_chrome_cookies()
    if not session:
        print("FAILED: Could not get cookies. Is Chrome open?")
        sys.exit(1)

    print("Fetching following list...")
    users = fetch_all_followings(session)
    print(f"\nTotal fetched: {len(users)}")

    if not users:
        print("No users fetched. Check auth.")
        sys.exit(1)

    print("Parsing data...")
    rows = [parse_user(u) for u in users]

    # Sort by followers descending, then mark top 800 as keep, rest as unfollow
    rows.sort(key=lambda r: -(r["Followers"] or 0))
    for i, r in enumerate(rows):
        r["Rank"] = i + 1
        r["Unfollow?"] = "keep" if i < 800 else "UNFOLLOW"

    print("Saving files...")
    save_excel(rows)
    save_csv(rows)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2)
    print(f"  Raw JSON saved: {OUTPUT_JSON}")

    to_unfollow = sum(1 for r in rows if r["Unfollow?"] == "UNFOLLOW")
    threshold = rows[799]["Followers"] if len(rows) >= 800 else "N/A"
    print(f"\n=== DONE ===")
    print(f"Total following:          {len(rows)}")
    print(f"Keep (top 800):           {len(rows) - to_unfollow}")
    print(f"Unfollow candidates:      {to_unfollow}")
    print(f"Follower threshold at #800: {threshold}")
    print(f"\nOpen: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
